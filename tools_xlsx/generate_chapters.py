#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Générateur de chapitres HTML à partir de fichiers Excel
Convertisseur de contenu Excel vers HTML interactif
"""

import openpyxl
import os
import sys
import markdown
import random
import hashlib
from datetime import UTC, datetime
from pathlib import Path
import json
import re



class ChapterGenerator:
    def __init__(self, parcours_slug=None):
        # Si un slug est fourni, les chapitres sont générés dans
        # parcours/src/{slug}/ (structure multi-parcours).
        # Sinon, on utilise le dossier legacy tools_xlsx/generated/.
        self.parcours_slug = parcours_slug
        if parcours_slug:
            self.output_dir = f"parcours/src/{parcours_slug}"
        else:
            self.output_dir = "tools_xlsx/generated"
        self.template_dir = "tools_xlsx/templates"

        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        Path(self.template_dir).mkdir(parents=True, exist_ok=True)

        self.create_base_template()
    
    def create_base_template(self):
        """Crée le template HTML de base — VERSION MULTI-PARCOURS"""
        template_content = (
            '<!DOCTYPE html>\n'
            '<html lang="fr">\n'
            '<head>\n'
            '    <meta charset="UTF-8">\n'
            '    <meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
            '    <title>{title}</title>\n'
            '    <link rel="icon" href="data:,">\n'
            '    <script>\n'
            '    (function() {\n'
            '      var base = "/";\n'
            '      if (window.location.pathname.indexOf("/cours-interactifs/") === 0) {\n'
            '        base = "/cours-interactifs/";\n'
            '      }\n'
            '      document.write(\'<base href="\' + base + \'">\');\n'
            '    })();\n'
            '    </script>\n'
            '    <script src="src/js/parcours.js"></script>\n'
            '    <script>\n'
            '    (function() {\n'
            '        if (!Parcours.token) {\n'
            '            window.location.replace(Parcours.loginUrl);\n'
            '        }\n'
            '    })();\n'
            '    </script>\n'
            '    <link rel="stylesheet" href="src/assets/css/style.css">\n'
            '    <script src="src/js/storage.js" defer></script>\n'
            '    <script src="src/js/dataStorage.js" defer></script>\n'
            '    <script src="src/js/main.js" defer></script>\n'
            '</head>\n'
            '<body class="chapter-page">\n'
            '    <div class="container">\n'
            '        <header class="chapter-header">\n'
            '            <h1>{title}</h1>\n'
            '            <div class="chapter-nav">\n'
            '                <button class="btn btn-secondary"\n'
            '                        onclick="window.location.href=Parcours.homeUrl">\n'
            '                    ← Retour au menu\n'
            '                </button>\n'
            '                {next_chapter_link}\n'
            '            </div>\n'
            '        </header>\n'
            '\n'
            '        <main class="chapter-content">\n'
            '            {content}\n'
            '        </main>\n'
            '\n'
            '        <footer class="chapter-footer">\n'
            '            <div class="progress-actions">\n'
            '                <button class="btn btn-primary"\n'
            '                        onclick="window.location.href=Parcours.homeUrl">\n'
            '                    Retour au menu\n'
            '                </button>\n'
            '                {next_chapter_button}\n'
            '            </div>\n'
            '        </footer>\n'
            '    </div>\n'
            '</body>\n'
            '</html>'
        )

        template_path = Path(self.template_dir) / "chapter_template.html"
        if not template_path.exists():
            with open(template_path, 'w', encoding='utf-8') as f:
                f.write(template_content)
    
    def convert_markdown_to_html(self, markdown_text):
        """Convertit du markdown enrichi en HTML"""
        if not markdown_text:
            return ""
        
        try:
            html = markdown.markdown(
                str(markdown_text),
                extensions=[
                    'extra',
                    'nl2br',
                    'sane_lists',
                    'tables',
                    'fenced_code',
                    'codehilite'
                ]
            )
            return html
        except Exception as e:
            print(f"⚠️ Erreur de conversion markdown: {e}")
            return str(markdown_text)
    
    def get_column_indices(self, headers):
        """Retourne un dictionnaire des indices de colonnes"""
        col_index = {}
        for idx, header in enumerate(headers):
            if header:
                header_str = str(header).strip().lower()
                # Normalisation
                header_str = header_str.replace('é', 'e').replace('è', 'e').replace('ê', 'e')
                header_str = header_str.replace('à', 'a').replace('â', 'a')
                header_str = header_str.replace('î', 'i').replace('ï', 'i')
                header_str = header_str.replace('ô', 'o').replace('ö', 'o')
                header_str = header_str.replace('û', 'u').replace('ü', 'u')
                header_str = header_str.replace('ç', 'c')
                header_str = header_str.replace('_', '')
                header_str = header_str.replace(' ', '')
                header_str = header_str.replace('(', '')
                header_str = header_str.replace(')', '')
                header_str = header_str.replace('-', '')
                header_str = header_str.replace('/', '')
                header_str = header_str.replace("'", '')
                
                col_index[header_str] = idx
        return col_index
    

    def generate_slug(self, title):
        """Génère un slug à partir d'un titre"""
        if not title:
            return ""
        # Enlever le préfixe "Chapitre x :"
        clean_title = re.sub(r'^Chapitre\s*\d+\s*:\s*', '', title, flags=re.IGNORECASE).strip()
        # Mettre en minuscules
        slug = clean_title.lower()
        # Remplacer les caractères accentués
        accents = {
            'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
            'à': 'a', 'â': 'a', 'ä': 'a',
            'î': 'i', 'ï': 'i',
            'ô': 'o', 'ö': 'o',
            'û': 'u', 'ü': 'u', 'ù': 'u',
            'ç': 'c',
            'œ': 'oe', 'æ': 'ae'
        }
        for accented, replacement in accents.items():
            slug = slug.replace(accented, replacement)
        # Remplacer les espaces et caractères spéciaux par des tirets
        slug = re.sub(r'[^a-z0-9]+', '-', slug)
        # Enlever les tirets multiples
        slug = re.sub(r'-+', '-', slug)
        # Enlever les tirets au début et à la fin
        slug = slug.strip('-')
        return slug

    def generate_chapters_index_json(self, chapters_metadata):
        """Génère un fichier JSON index pour les chapitres"""
        if not chapters_metadata:
            chapters_metadata = []
        
        chapters_list = []

        for metadata in chapters_metadata:
            chapter_entry = {
                "id": metadata["id"],
                "slug": metadata["slug"],
                "title": metadata["title"],
                "href": metadata["href"],
                "sheetName": metadata["sheetName"],
                "chapterHash": metadata["chapterHash"],

                # Questions
                "questionCount": metadata["questionCount"],

                # Cours
                "courseCount": metadata.get("courseCount", 0),
                "courseValidationCount": metadata.get("courseValidationCount", 0),
                "courses": metadata.get("courses", []),

                # Progression
                "progressItemCount": metadata.get(
                    "progressItemCount",
                    metadata["questionCount"]
                ),
                "progressBreakdown": metadata.get(
                    "progressBreakdown",
                    {
                        "questions": metadata["questionCount"],
                        "coursesToValidate": 0
                    }
                ),

                # Divers
                "maxPoints": metadata["maxPoints"],
                "estimatedDuration": metadata["estimatedDuration"],
                "hasValidationSteps": metadata["hasValidationSteps"],
                "hasManualCorrection": metadata["hasManualCorrection"],
                "types": metadata["types"],
                "questions": metadata["questions"]
            }
            chapters_list.append(chapter_entry)
        
        # Trier par id
        chapters_list.sort(key=lambda x: x['id'])

        # Calculer les totaux globaux
        total_questions = sum(ch.get("questionCount", 0) for ch in chapters_list)
        total_course_validations = sum(ch.get("courseValidationCount", 0) for ch in chapters_list)
        total_progress_items = sum(ch.get("progressItemCount", ch.get("questionCount", 0)) for ch in chapters_list)

        # Calculer le contentHash = md5(chapterHashes)
        chapter_hashes = "".join(ch.get("chapterHash", "") for ch in chapters_list)
        content_hash = hashlib.md5(chapter_hashes.encode('utf-8')).hexdigest()[:10]

        # Générer les timestamps
        now = datetime.now(UTC).isoformat()

        # Créer la structure finale
        index_data = {
            "version": now,
            "generatedAt": now,
            "contentHash": content_hash,
            "totalChapters": len(chapters_list),

            # Existant
            "totalQuestions": total_questions,

            # Nouveau : progression globale
            "totalCourseValidations": total_course_validations,
            "totalProgressItems": total_progress_items,

            "chapters": chapters_list
        }

        # Sauvegarder le JSON
        json_path = Path(self.output_dir) / "chapters_index.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(index_data, f, ensure_ascii=False, indent=4)
        
        print(f"✅ Fichier JSON généré : {json_path}")
        return str(json_path)
    
    def generate_from_excel(self, excel_file):
        """Génère tous les chapitres à partir d'un fichier Excel"""
        try:
            print(f"🔍 Lecture du fichier Excel : {excel_file}")
            
            if not os.path.exists(excel_file):
                raise FileNotFoundError(f"Le fichier {excel_file} n'existe pas")
            
            workbook = openpyxl.load_workbook(excel_file)
            print(f"✅ Classeur chargé avec {len(workbook.sheetnames)} feuilles")
            
            generated_files = []
            chapters_metadata = []
            total_questions = 0
            total_points = 0
            
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                worksheet = workbook[sheet_name]
                result, metadata = self.generate_chapter(worksheet, i, sheet_name)
                if result:
                    generated_files.append(result)
                    chapters_metadata.append(metadata)
                    total_questions += metadata.get("questionCount", 0)
                    total_points += metadata.get("maxPoints", 0)
            
            print(f"✅ Génération terminée : {len(generated_files)} fichiers générés")
            # Générer JSON
            self.generate_chapters_index_json(chapters_metadata)

            return {
                'success': True,
                'message': f'{len(generated_files)} chapitres générés avec succès',
                'files': generated_files,
                'chapters': chapters_metadata,
                'totalQuestions': total_questions,
                'totalPoints': total_points
            }
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération : {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _get_col(self, row, col_index, *keys, default=""):
        """Récupère une valeur de cellule en essayant plusieurs clés de colonne"""
        for key in keys:
            idx = col_index.get(key)
            if idx is not None and idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                if val:
                    return val
        return default

    def generate_chapter(self, worksheet, chapter_number, chapter_title):
        """Génère un chapitre HTML à partir d'une feuille Excel"""
        try:
            print(f"📝 Génération du chapitre {chapter_number} : {chapter_title}")
            
            rows = list(worksheet.iter_rows(values_only=True))
            if len(rows) < 2:
                print(f"⚠️  Feuille {chapter_title} vide, ignorée")
                return None, {}
            
            headers = rows[0]
            col_index = self.get_column_indices(headers)
            print(f"📋 Colonnes trouvées : {list(col_index.keys())}")
            
            data_rows = rows[1:]
            
            html_content = ""
            question_count = 0
            max_points = 0
            course_count = 0
            course_validation_count = 0
            has_validation_steps = False
            has_manual_correction = False
            types_count = {
                "cours": 0,
                "qcm": 0,
                "ouverte": 0,
                "courte": 0,
                "selection": 0
            }
            questions_list = []
            courses_list = []
            
            for row in data_rows:
                if len(row) < 2:
                    continue
                
                # Récupérer les valeurs par nom de colonne (ancien + nouveau format)
                content_type = self._get_col(row, col_index, 'type', default="")
                content = self._get_col(row, col_index, 'contenu', 'enonce', default="")
                
                if not content_type or not content:
                    continue
                
                content_type = str(content_type).lower()
                
                if content_type == 'cours':
                    types_count["cours"] += 1
                    course_count += 1

                    # Vérifier si le cours nécessite une validation utilisateur
                    regle = self._get_col(row, col_index, 'regle', default="")
                    requires_validation = bool(
                        regle and 'validation' in str(regle).lower()
                    )

                    if requires_validation:
                        has_validation_steps = True
                        course_validation_count += 1

                    # Enregistrer les informations détaillées du cours
                    courses_list.append({
                        "index": course_count - 1,
                        "requiresValidation": requires_validation
                    })

                    html_content += self.generate_course_content(row, col_index, content)
                elif content_type == 'qcm':
                    question_count += 1
                    types_count["qcm"] += 1
                    html_part, question_meta = self.generate_qcm_content(row, col_index, question_count, chapter_number)
                    if question_meta:
                        questions_list.append(question_meta)
                        max_points += question_meta.get("points", 0)
                        if question_meta.get("correctionType") in ('manuel', 'semi'):
                            has_manual_correction = True
                    html_content += html_part
                elif content_type == 'ouverte':
                    question_count += 1
                    types_count["ouverte"] += 1
                    html_part, question_meta = self.generate_open_content(row, col_index, question_count, chapter_number)
                    if question_meta:
                        questions_list.append(question_meta)
                        max_points += question_meta.get("points", 0)
                        if question_meta.get("correctionType") in ('manuel', 'semi'):
                            has_manual_correction = True
                    html_content += html_part
                elif content_type == 'courte':
                    question_count += 1
                    types_count["courte"] += 1
                    html_part, question_meta = self.generate_short_content(row, col_index, question_count, chapter_number)
                    if question_meta:
                        questions_list.append(question_meta)
                        max_points += question_meta.get("points", 0)
                        if question_meta.get("correctionType") in ('manuel', 'semi'):
                            has_manual_correction = True
                    html_content += html_part
                elif content_type == 'selection':
                    question_count += 1
                    types_count["selection"] += 1
                    html_part, question_meta = self.generate_selection_content(row, col_index, question_count, chapter_number)
                    if question_meta:
                        questions_list.append(question_meta)
                        max_points += question_meta.get("points", 0)
                        if question_meta.get("correctionType") in ('manuel', 'semi'):
                            has_manual_correction = True
                    html_content += html_part
            
            # Ajouter le bouton de validation globale
            html_content += '''
                <div class="global-validation hidden">
                    <button class="btn btn-primary" onclick="window.studentWorkEditor.validateAllQuestions()">
                        ✅ Valider toutes les réponses
                    </button>
                    <div id="global-feedback" class="feedback"></div>
                </div>
            '''
            
            # Générer les liens de navigation
            next_chapter_link = ""
            next_chapter_button = ""
            
            if chapter_number < len(worksheet.parent.sheetnames):
                next_chapter_num = chapter_number + 1
                next_chapter_link = f'<a href="chapitre{next_chapter_num}.html" class="btn btn-primary">Chapitre {next_chapter_num} →</a>'
                next_chapter_button = f'<button class="btn btn-primary" onclick="window.location.href=\'chapitre{next_chapter_num}.html\'">Chapitre {next_chapter_num} →</button>'
            
            # Charger le template
            template_path = Path(self.template_dir) / "chapter_template.html"
            with open(template_path, 'r', encoding='utf-8') as f:
                template = f.read()
            
            # Utiliser str.replace() pour éviter les conflits avec les { } dans le HTML
            final_html = template
            final_html = final_html.replace('{title}', f"Chapitre {chapter_number} : {chapter_title}")
            final_html = final_html.replace('{content}', html_content)
            final_html = final_html.replace('{next_chapter_link}', next_chapter_link)
            final_html = final_html.replace('{next_chapter_button}', next_chapter_button)
            
            # Sauvegarder
            output_file = Path(self.output_dir) / f"chapitre{chapter_number}.html"
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(final_html)
            
            # Nettoyer le titre
            clean_title = re.sub(r'^Chapitre\s*\d+\s*:\s*', '', chapter_title, flags=re.IGNORECASE).strip()
            slug = self.generate_slug(clean_title)
            
            # Calculer le chapterHash = md5(title + questionHashes)
            question_hashes = "".join(q.get("questionHash", "") for q in questions_list)
            chapter_hash_content = f"{clean_title}{question_hashes}"
            chapter_hash = hashlib.md5(chapter_hash_content.encode('utf-8')).hexdigest()[:10]
            
            # Calculer estimatedDuration
            estimated_duration = round(question_count * 2 + course_count)
            
            # Créer les métadonnées du chapitre
            metadata = {
                "id": chapter_number,
                "slug": slug,
                "title": clean_title,
                "href": f"chapitre{chapter_number}.html",
                "sheetName": chapter_title,
                "chapterHash": chapter_hash,

                # Questions
                "questionCount": question_count,

                # Cours
                "courseCount": course_count,
                "courseValidationCount": course_validation_count,
                "courses": courses_list,

                # Progression globale
                "progressItemCount": question_count + course_validation_count,
                "progressBreakdown": {
                    "questions": question_count,
                    "coursesToValidate": course_validation_count
                },

                # Divers
                "maxPoints": max_points,
                "estimatedDuration": estimated_duration,
                "hasValidationSteps": has_validation_steps,
                "hasManualCorrection": has_manual_correction,
                "types": types_count,
                "questions": questions_list
            }
            
            print(f"✅ Chapitre {chapter_number} généré : {output_file}")
            return str(output_file), metadata
            
        except Exception as e:
            print(f"❌ Erreur lors de la génération du chapitre {chapter_number} : {str(e)}")
            return None, {}
    
    def _resolve_correct_indices(self, correct_values_str, choice_list):
        """
        Convertit les valeurs textuelles des bonnes réponses (séparées par \n)
        en indices 0-based dans la liste des choix.
        Exemple: "ok\nko" avec choice_list=["ok","ko"] → [0, 1]
        """
        if not correct_values_str or not choice_list:
            return []
        
        indices = []
        for val in str(correct_values_str).split('\n'):
            val = val.strip()
            if val and val in choice_list:
                indices.append(choice_list.index(val))
        return indices

    def generate_course_content(self, row, col_index, content):
        """Génère le contenu de type cours"""
        if not content:
            return ""
        
        regle = self._get_col(row, col_index, 'regle', default="")
        html_content = self.convert_markdown_to_html(content)
        
        # Si c'est un cours avec validation (bouton "J'ai lu")
        if regle and 'validation' in str(regle).lower():
            return f'''
            <section class="course-content">
                <div class="content-box">
                    {html_content}
                    <div class="course-validation" style="margin-top: 1rem; text-align: right;">
                        <button class="btn btn-secondary" onclick="window.studentWorkEditor.validateCourse(this)">
                            ✅ J'ai lu et compris
                        </button>
                    </div>
                </div>
            </section>
            '''
        else:
            return f'''
            <section class="course-content">
                <div class="content-box">
                    {html_content}
                </div>
            </section>
            '''
        
    def generate_qcm_content(self, row, col_index, question_index, chapter_number):
        """Génère le contenu de type QCM"""
        # Récupérer les valeurs (ancien + nouveau format)
        question_text = self._get_col(row, col_index, 'contenu', 'enonce', default="")
        hint = self._get_col(row, col_index, 'indication', 'indiceaide', default="")
        points_str = self._get_col(row, col_index, 'points', default="1")
        try:
            points = int(float(points_str))
        except:
            points = 1
        choices = self._get_col(row, col_index, 'choix', 'propositionsdereponse', default="")
        correct_answers_text = self._get_col(row, col_index, 'choix_corrects', 'bonnesreponses', default="")
        correction_type = self._get_col(row, col_index, 'correction', default="auto")
        regle = self._get_col(row, col_index, 'regle', default="")
        order = self._get_col(row, col_index, 'ordre_choix', 'ordredesreponses', default="")
        
        if not question_text or not choices:
            return "", {}
        
        question_id = f"ch{chapter_number}_q{question_index}"
        
        # Parser les choix
        choice_list = [choice.strip() for choice in str(choices).split('\n') if choice.strip()]
        if not choice_list:
            return "", {}
        
        # Résoudre les indices des bonnes réponses à partir des valeurs textuelles
        correct_indices_0based = self._resolve_correct_indices(correct_answers_text, choice_list)
        
        # Mélanger si ordre aléatoire
        if order and 'aleatoire' in str(order).lower():
            correct_values = []
            for idx in correct_indices_0based:
                if idx < len(choice_list):
                    correct_values.append(choice_list[idx])
            random.shuffle(choice_list)
            correct_indices_0based = []
            for val in correct_values:
                if val in choice_list:
                    correct_indices_0based.append(choice_list.index(val))
        
        # Déterminer si c'est unique ou multiple
        is_multiple = 'multiple' in str(regle).lower()
        input_type = "checkbox" if is_multiple else "radio"

        # Sérialiser les bonnes réponses en JSON pour data-correct-answers
        correct_answers_json = json.dumps(correct_indices_0based)
        
        # Générer les options HTML
        choices_html = ""
        for i, choice in enumerate(choice_list):
            choices_html += f'''
                                    <div class="choice-option">
                                        <input type="{input_type}" name="qcm_{question_id}" value="{i}" id="qcm_{question_id}_{i}">
                                        <label for="qcm_{question_id}_{i}">{choice}</label>
                                    </div>
                                '''
        
        html = f'''
                            <section class="question-section" data-question-id="{question_id}" 
                                    data-correction-type="{correction_type}" 
                                    data-points="{points}"
                                    data-correct-answers='{correct_answers_json}'>
                                <div class="question-box">
                                    <div class="question-header">
                                        <div class="question-title">
                                            <h3>Question {question_index}</h3>
                                        </div>
                                        <div class="question-meta">
                                            <span class="points-badge">⭐ {points} point{'s' if points > 1 else ''}</span>
                                            {self.generate_hint_badge(hint, question_id)}
                                            <span class="correction-badge correction-{correction_type}">{self.get_correction_label(correction_type)}</span>
                                        </div>
                                    </div>
                                    <div class="question-text">{self.convert_markdown_to_html(question_text)}</div>
                                    {self.generate_hint_content(hint, question_id)}
                                    <div class="choices">
                                        {choices_html}
                                    </div>
                                    <div class="question-actions">
                                        <button class="btn-check-answer" onclick="window.studentWorkEditor.handleAnswer('{question_id}', '{correction_type}', {points})">
                                            {self.get_button_label(correction_type)}
                                        </button>
                                        <div class="feedback" id="feedback_{question_id}"></div>
                                    </div>
                                </div>
                            </section>
                            '''
        
        # Créer les métadonnées de la question
        hash_content = f"{question_text}{'qcm'}{correct_indices_0based}{points}"
        question_hash = hashlib.md5(hash_content.encode('utf-8')).hexdigest()[:10]
        
        metadata = {
            "id": question_id,
            "index": question_index,
            "type": "qcm",
            "title": f"Question {question_index}",
            "questionText": str(question_text),
            "points": points,
            "correctionType": str(correction_type),
            "rule": str(regle) if regle else None,
            "required": True,
            "hasHint": bool(hint),
            "questionHash": question_hash,
            "choiceCount": len(choice_list),
            "allowMultiple": is_multiple,
            "correctAnswers": correct_indices_0based,
            "options": choice_list,
            "minLength": None,
            "maxAttempts": None
        }
        
        return html, metadata


    def generate_short_content(self, row, col_index, question_index, chapter_number):
        """Génère le contenu de type réponse courte"""
        question_text = self._get_col(row, col_index, 'contenu', 'enonce', default="")
        hint = self._get_col(row, col_index, 'indication', 'indiceaide', default="")
        points_str = self._get_col(row, col_index, 'points', default="1")
        try:
            points = int(float(points_str))
        except:
            points = 1
        correct_answers = self._get_col(row, col_index, 'choix_corrects', 'bonnesreponses', default="")
        correction_type = self._get_col(row, col_index, 'correction', default="auto")
        regle = self._get_col(row, col_index, 'regle', default="")
        
        input_type = "number" if 'nombre' in str(regle).lower() else "text"
        
        question_id = f"ch{chapter_number}_q{question_index}"
        
        # Nettoyer les bonnes réponses
        correct_answers_list = []
        if correct_answers:
            for ans in str(correct_answers).split('\n'):
                ans = ans.strip()
                if ans:
                    correct_answers_list.append(ans.lower())

        # Sérialiser en JSON pour data-correct-answers
        correct_answers_json = json.dumps(correct_answers_list)
        
        html = f'''
                                <section class="question-section" data-question-id="{question_id}" 
                                        data-correction-type="{correction_type}" 
                                        data-points="{points}" 
                                        data-correct-answers='{correct_answers_json}'>
                                    <div class="question-box">
                                    <div class="question-header">
                                        <div class="question-title">
                                            <h3>Question {question_index}</h3>
                                        </div>
                                        <div class="question-meta">
                                            <span class="points-badge">⭐ {points} point{'s' if points > 1 else ''}</span>
                                            {self.generate_hint_badge(hint, question_id)}
                                            <span class="correction-badge correction-{correction_type}">{self.get_correction_label(correction_type)}</span>
                                        </div>
                                    </div>
                                    <div class="question-text">{self.convert_markdown_to_html(question_text)}</div>
                                    {self.generate_hint_content(hint, question_id)}
                                    <div class="answer-area">
                                        <input type="{input_type}" id="short_{question_id}" placeholder="Votre réponse...">
                                    </div>
                                    <div class="question-actions">
                                        <button class="btn-check-answer" onclick="window.studentWorkEditor.handleAnswer('{question_id}', '{correction_type}', {points})">
                                            {self.get_button_label(correction_type)}
                                        </button>
                                        <div class="feedback" id="feedback_{question_id}"></div>
                                    </div>
                                    </div>
                                </section>
                                '''
        
        # Créer les métadonnées de la question
        hash_content = f"{question_text}{'courte'}{correct_answers_list}{points}"
        question_hash = hashlib.md5(hash_content.encode('utf-8')).hexdigest()[:10]
        
        metadata = {
            "id": question_id,
            "index": question_index,
            "type": "courte",
            "title": f"Question {question_index}",
            "questionText": str(question_text),
            "points": points,
            "correctionType": str(correction_type),
            "rule": str(regle) if regle else None,
            "required": True,
            "hasHint": bool(hint),
            "questionHash": question_hash,
            "choiceCount": None,
            "allowMultiple": False,
            "correctAnswers": correct_answers_list,
            "minLength": None,
            "maxAttempts": None
        }
        
        return html, metadata


    def generate_open_content(self, row, col_index, question_index, chapter_number):
        """Génère le contenu de type réponse ouverte avec validation de longueur"""
        question_text = self._get_col(row, col_index, 'contenu', 'enonce', default="")
        hint = self._get_col(row, col_index, 'indication', 'indiceaide', default="")
        points_str = self._get_col(row, col_index, 'points', default="1")
        try:
            points = int(float(points_str))
        except:
            points = 1
        correction_type = self._get_col(row, col_index, 'correction', default="semi")
        regle = self._get_col(row, col_index, 'regle', default="")
        
        min_length = 0
        if 'texte(' in str(regle):
            try:
                min_length = int(str(regle).split('(')[1].split(')')[0])
            except:
                min_length = 0
        
        question_id = f"ch{chapter_number}_q{question_index}"
        
        html = f'''
                        <section class="question-section" data-question-id="{question_id}" data-correction-type="{correction_type}" data-points="{points}" data-min-length="{min_length}">
                            <div class="question-box">
                                <div class="question-header">
                                    <div class="question-title">
                                        <h3>Question {question_index}</h3>
                                    </div>
                                    <div class="question-meta">
                                        <span class="points-badge">⭐ {points} point{'s' if points > 1 else ''}</span>
                                        {self.generate_hint_badge(hint, question_id)}
                                        <span class="correction-badge correction-{correction_type}">{self.get_correction_label(correction_type)}</span>
                                    </div>
                                </div>
                                <div class="question-text">{self.convert_markdown_to_html(question_text)}</div>
                                {self.generate_hint_content(hint, question_id)}
                                <div class="answer-area">
                                    <textarea id="{question_id}" placeholder="Votre réponse..." rows="4" data-min-length="{min_length}"></textarea>
                                    {f'<small style="color: #666; display: block; margin-top: 0.25rem;">Minimum {min_length} caractères</small>' if min_length > 0 else ''}
                                </div>
                                <div class="question-actions">
                                    <button class="btn-check-answer" onclick="window.studentWorkEditor.handleOpenAnswer('{question_id}', '{correction_type}', {points}, {min_length})">
                                        {self.get_button_label(correction_type)}
                                    </button>
                                    <div class="feedback" id="feedback_{question_id}"></div>
                                </div>
                            </div>
                        </section>
                        '''
        
        # Créer les métadonnées de la question
        hash_content = f"{question_text}{'ouverte'}[]{points}"
        question_hash = hashlib.md5(hash_content.encode('utf-8')).hexdigest()[:10]
        
        metadata = {
            "id": question_id,
            "index": question_index,
            "type": "ouverte",
            "title": f"Question {question_index}",
            "questionText": str(question_text),
            "points": points,
            "correctionType": str(correction_type),
            "rule": str(regle) if regle else None,
            "required": True,
            "hasHint": bool(hint),
            "questionHash": question_hash,
            "choiceCount": None,
            "allowMultiple": False,
            "correctAnswers": [],
            "minLength": min_length if min_length > 0 else None,
            "maxAttempts": None
        }
        
        return html, metadata

    def generate_selection_content(self, row, col_index, question_index, chapter_number):
        """Génère le contenu de type sélection (liste déroulante)"""
        question_text = self._get_col(row, col_index, 'contenu', 'enonce', default="")
        hint = self._get_col(row, col_index, 'indication', 'indiceaide', default="")
        points_str = self._get_col(row, col_index, 'points', default="1")
        try:
            points = int(float(points_str))
        except:
            points = 1
        choices = self._get_col(row, col_index, 'choix', 'propositionsdereponse', default="")
        correct_answers_text = self._get_col(row, col_index, 'choix_corrects', 'bonnesreponses', default="")
        correction_type = self._get_col(row, col_index, 'correction', default="auto")
        
        if not question_text or not choices:
            return "", {}
        
        question_id = f"ch{chapter_number}_q{question_index}"
        
        # Parser les choix
        choice_list = [choice.strip() for choice in str(choices).split('\n') if choice.strip()]
        if not choice_list:
            return "", {}
        
        # Résoudre l'index de la bonne réponse à partir de la valeur textuelle
        correct_indices = self._resolve_correct_indices(correct_answers_text, choice_list)
        correct_index = correct_indices[0] if correct_indices else 0

        # Sérialiser en JSON pour data-correct-answers
        correct_answers_json = json.dumps([correct_index])
        
        # Générer la liste déroulante
        options_html = '<option value="">-- Choisissez une réponse --</option>'
        for i, choice in enumerate(choice_list):
            options_html += f'<option value="{i}">{choice}</option>'
        
        html = f'''
                    <section class="question-section" data-question-id="{question_id}" 
                            data-correction-type="{correction_type}" 
                            data-points="{points}"
                            data-correct-answers='{correct_answers_json}'>
                        <div class="question-box">
                        <div class="question-header">
                            <div class="question-title">
                                <h3>Question {question_index}</h3>
                            </div>
                            <div class="question-meta">
                                <span class="points-badge">⭐ {points} point{'s' if points > 1 else ''}</span>
                                {self.generate_hint_badge(hint, question_id)}
                                <span class="correction-badge correction-{correction_type}">{self.get_correction_label(correction_type)}</span>
                            </div>
                        </div>
                        <div class="question-text">{self.convert_markdown_to_html(question_text)}</div>
                        {self.generate_hint_content(hint, question_id)}
                        <div class="answer-area">
                            <select id="{question_id}" class="select-answer">
                                {options_html}
                            </select>
                        </div>
                        <div class="question-actions">
                            <button class="btn-check-answer" onclick="window.studentWorkEditor.handleAnswer('{question_id}', '{correction_type}', {points})">
                                {self.get_button_label(correction_type)}
                            </button>
                            <div class="feedback" id="feedback_{question_id}"></div>
                        </div>
                        </div>
                    </section>
                    '''
        
        # Créer les métadonnées de la question
        hash_content = f"{question_text}{'selection'}{[correct_index]}{points}"
        question_hash = hashlib.md5(hash_content.encode('utf-8')).hexdigest()[:10]
        
        metadata = {
            "id": question_id,
            "index": question_index,
            "type": "selection",
            "title": f"Question {question_index}",
            "questionText": str(question_text),
            "points": points,
            "correctionType": str(correction_type),
            "rule": None,
            "required": True,
            "hasHint": bool(hint),
            "questionHash": question_hash,
            "choiceCount": len(choice_list),
            "allowMultiple": False,
            "correctAnswers": [correct_index],
            "options": choice_list,
            "minLength": None,
            "maxAttempts": None
        }
        
        return html, metadata
        
    def get_correction_label(self, correction_type):
        """Retourne le libellé du type de correction"""
        labels = {
            'auto': '🔍 Auto',
            'semi': '⚡ Semi-auto',
            'manuel': '📝 Correction manuelle'
        }
        return labels.get(correction_type, '🔍 Auto')

    def get_button_label(self, correction_type):
        """Retourne le texte du bouton selon le type de correction"""
        labels = {
            'auto': '✓ Vérifier',
            'semi': '✓ Vérifier',
            'manuel': '📌 Envoyer au formateur'
        }
        return labels.get(correction_type, '✓ Vérifier')
    
    def generate_hint_badge(self, hint, question_id):
        """Génère un badge pour l'indication"""
        if not hint:
            return ""
        return f'''
            <button class="hint-badge" data-hint-btn onclick="window.studentWorkEditor.toggleHint('hint_{question_id}')" type="button">
                💡 Indication
            </button>
        '''

    def generate_hint_content(self, hint, question_id):
        """Génère le contenu de l'indication (caché par défaut)"""
        if not hint:
            return ""
        return f'''
            <div class="hint-container" id="hint_{question_id}" style="display: none;">
                <div class="hint-content">
                    {self.convert_markdown_to_html(str(hint))}
                </div>
            </div>
        '''

def _update_parcours_registry(slug, label=None):
    """
    Met à jour parcours/parcours.json avec le nouveau parcours.
    Crée le fichier s'il n'existe pas.
    N'écrase pas les entrées existantes.
    Le label est déduit du slug si non fourni (ex: "nsi-term" → "NSI - Term").
    """
    registry_path = Path("parcours/parcours.json")
    registry_path.parent.mkdir(parents=True, exist_ok=True)

    # Charger l'existant
    existing = []
    if registry_path.exists():
        try:
            with open(registry_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        except Exception:
            existing = []

    # Vérifier si le slug est déjà présent
    if any(p.get('slug') == slug for p in existing):
        print(f"   ℹ️  Parcours '{slug}' déjà dans parcours/parcours.json")
        return

    # Générer un label lisible depuis le slug si non fourni
    if not label:
        label = slug.replace('-', ' ').replace('_', ' ').title()

    existing.append({"slug": slug, "label": label})

    with open(registry_path, 'w', encoding='utf-8') as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"   ✅ parcours/parcours.json mis à jour → ajouté '{slug}' ({label})")
    print(f"   💡 Pour personnaliser le label, éditez parcours/parcours.json")



def main():
    """Fonction principale pour l'exécution en ligne de commande"""
    import argparse
    parser = argparse.ArgumentParser(description='Générateur de chapitres HTML depuis Excel')
    parser.add_argument('excel_file', help='Fichier Excel source')
    parser.add_argument('--parcours', '-p',
                        help='Slug du parcours (ex: nsi-term). Génère dans parcours/src/{slug}/',
                        default=None)
    args = parser.parse_args()

    excel_file = args.excel_file
    slug       = args.parcours

    if not os.path.exists(excel_file):
        print(f"❌ Erreur: Le fichier {excel_file} n'existe pas")
        sys.exit(1)

    if slug:
        print(f"🎯 Parcours ciblé : {slug}")
        output = os.path.abspath(f'parcours/src/{slug}')
    else:
        output = os.path.abspath('tools_xlsx/generated')

    print("🚀 Démarrage de la génération des chapitres...")
    print(f"📁 Répertoire de sortie: {output}")

    generator = ChapterGenerator(parcours_slug=slug)
    result = generator.generate_from_excel(excel_file)
    
    if result['success']:
        print(f"\n✅ Génération réussie !")
        print(f"📄 {result['message']}")
        print(f"\n📋 Fichiers générés:")
        for file_path in result['files']:
            print(f"   - {file_path}")

        # Mise à jour automatique de parcours/parcours.json
        if slug:
            _update_parcours_registry(slug)

        print(f"\n💡 Instructions:")
        if slug:
            print(f"   Fichiers disponibles dans parcours/src/{slug}/")
        else:
            print(f"   Copiez les fichiers depuis tools_xlsx/generated/ vers le dossier  du parcours")
            print(f"   Exemple: cp tools_xlsx/generated/*.html parcours/src/nsi-term/")
    else:
        print(f"\n❌ Erreur lors de la génération: {result['error']}")
        sys.exit(1)

if __name__ == "__main__":
    main()